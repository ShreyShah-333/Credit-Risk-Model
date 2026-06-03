# ============================================================
# app.py — Credit Risk Model Web App
# Run with: python app.py
# Opens at: http://127.0.0.1:8080
# ============================================================

from flask import Flask, render_template_string, request, jsonify
import anthropic
import json
import math
import re
import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import gspread
import sys
from google.oauth2.service_account import Credentials
from datetime import datetime

app = Flask(__name__)

# ── CONFIG ────────────────────────────────────────────────────
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
MODEL             = "claude-sonnet-4-5"
SHEET_ID          = "13pFG_57Eb6EOGeuw90B65Z6q2LkuhSqfAFLPYE83zrk"
EXCEL_PATH        = "/tmp/Credit_Risk_Portfolio.xlsx"

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

CONFIG = {
    "thresholds": {
        "dscr_min":1.25,"leverage_max":0.65,"interest_cov_min":3.0,
        "current_ratio_min":1.5,"quick_ratio_min":1.0,
        "dti_max":0.40,"utilisation_max":0.80,
    },
    "lgd_by_collateral": {
        "Real Estate":0.35,"Equipment":0.55,"Receivables":0.60,
        "Inventory":0.70,"Unsecured":0.90,"Government Bond":0.10,"Cash Deposit":0.05,
    },
    "ccf_by_type": {
        "Term Loan":1.00,"Revolving Credit":0.75,
        "Line of Credit":0.75,"Overdraft":0.80,"Mortgage":1.00,
    },
    "sector_base_pd": {
        "Technology":0.025,"Healthcare":0.018,"Real Estate":0.032,
        "Manufacturing":0.028,"Retail":0.038,"Energy":0.035,
        "Financial Services":0.020,"Construction":0.042,
        "Hospitality":0.055,"Agriculture":0.030,"Other":0.030,
    },
}

# ============================================================
# HTML TEMPLATE
# ============================================================
HTML = """
<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Credit Risk Model AI Agent</title>
<style>
  * { box-sizing:border-box; margin:0; padding:0; }
  body { font-family:'Segoe UI',Arial,sans-serif; background:#F8FAFC; color:#1F2937; }
  .hero {
    background:linear-gradient(135deg,#0A1628 0%,#1A3A5C 100%);
    padding:32px 40px; color:white; text-align:center;
  }
  .hero h1 { font-size:28px; font-weight:800; letter-spacing:1px; }
  .hero p  { color:#F4A261; font-size:13px; margin-top:6px; }
  .tags    { margin-top:14px; display:flex; justify-content:center; gap:8px; flex-wrap:wrap; }
  .tag     { background:rgba(255,255,255,0.15); color:white; padding:4px 14px;
             border-radius:20px; font-size:11px; font-weight:600; }
  .container { max-width:1100px; margin:30px auto; padding:0 20px; }
  .card {
    background:white; border:1px solid #E2E8F0; border-radius:12px;
    overflow:hidden; box-shadow:0 2px 8px rgba(0,0,0,0.05); margin-bottom:20px;
  }
  .card-header {
    background:#0A1628; color:white; padding:12px 18px;
    font-weight:700; font-size:13px;
  }
  .card-body { padding:20px; }
  .form-row   { display:grid; grid-template-columns:1fr 1fr; gap:14px; margin-bottom:14px; }
  .form-group { display:flex; flex-direction:column; gap:5px; }
  .form-group label {
    font-size:11px; font-weight:600; color:#6B7280;
    text-transform:uppercase; letter-spacing:0.6px;
  }
  .form-group input,
  .form-group select {
    padding:9px 12px; border:1.5px solid #E2E8F0; border-radius:8px;
    font-size:13px; color:#1F2937; outline:none; transition:border 0.2s;
  }
  .form-group input:focus,
  .form-group select:focus { border-color:#2E86AB; }
  .toggle-row { display:flex; gap:10px; margin-bottom:14px; }
  .toggle-btn {
    flex:1; padding:10px; border:2px solid #E2E8F0; border-radius:8px;
    background:white; font-size:13px; font-weight:600; cursor:pointer;
    color:#6B7280; transition:all 0.2s;
  }
  .toggle-btn.active { border-color:#2E86AB; background:#EEF2FF; color:#2E86AB; }
  .submit-btn {
    width:100%; padding:15px; background:linear-gradient(135deg,#0A1628,#1A3A5C);
    color:white; border:none; border-radius:10px; font-size:15px;
    font-weight:700; cursor:pointer; margin-top:10px; transition:opacity 0.2s;
  }
  .submit-btn:hover   { opacity:0.9; }
  .submit-btn:disabled{ opacity:0.6; cursor:not-allowed; }
  .progress { display:none; background:white; border:1px solid #E2E8F0;
              border-radius:12px; padding:20px; margin:20px 0; }
  .progress-title { font-weight:700; color:#0A1628; margin-bottom:14px; font-size:14px; }
  .step {
    display:flex; align-items:center; gap:10px; padding:8px 12px;
    border-radius:8px; margin:4px 0; font-size:13px; background:#F8FAFC;
    border-left:3px solid #E2E8F0; transition:all 0.3s;
  }
  .step.active  { border-left-color:#2E86AB; background:#EEF2FF; color:#1A3A5C; }
  .step.done    { border-left-color:#10B981; background:#F0FDF4; color:#065F46; }
  .step.error   { border-left-color:#EF4444; background:#FEE2E2; color:#991B1B; }
  .step-icon    { font-size:16px; width:24px; text-align:center; }
  #results { display:none; }
  .verdict {
    background:linear-gradient(135deg,#0A1628,#1A3A5C);
    border-radius:12px; padding:28px; text-align:center; color:white; margin:20px 0;
  }
  .verdict-label { font-size:11px; color:#F4A261; letter-spacing:2px;
                   text-transform:uppercase; margin-bottom:8px; }
  .verdict-score { font-size:64px; font-weight:900; line-height:1; }
  .verdict-sub   { font-size:13px; color:rgba(255,255,255,0.6); margin-top:4px; }
  .verdict-pills { display:flex; justify-content:center; gap:10px; margin-top:16px; flex-wrap:wrap; }
  .pill { padding:7px 22px; border-radius:20px; font-size:13px; font-weight:700; color:white; }
  .metrics-grid { display:grid; grid-template-columns:repeat(4,1fr); gap:12px; margin:16px 0; }
  .metric {
    background:white; border:1px solid #E2E8F0; border-radius:10px;
    padding:14px; text-align:center; border-top:4px solid #2E86AB;
  }
  .metric-label { font-size:10px; color:#6B7280; text-transform:uppercase;
                  letter-spacing:0.8px; margin-bottom:4px; }
  .metric-value { font-size:20px; font-weight:700; color:#0A1628; }
  .rationale-box {
    background:#EEF2FF; border-left:4px solid #2E86AB; padding:14px 18px;
    border-radius:0 8px 8px 0; font-size:13px; line-height:1.7; color:#1F2937; margin:12px 0;
  }
  .reasoning-box {
    background:#F8F0FF; border-left:4px solid #8B5CF6; padding:14px 18px;
    border-radius:0 8px 8px 0; font-family:'Courier New',monospace;
    font-size:11px; color:#0A1628; margin:12px 0; line-height:1.8;
  }
  .two-col { display:grid; grid-template-columns:1fr 1fr; gap:12px; margin:12px 0; }
  .strengths-box { background:#F0FDF4; border-radius:8px; padding:14px 18px; }
  .flags-box     { background:#FEF2F2; border-radius:8px; padding:14px 18px; }
  .box-title { font-weight:700; font-size:13px; margin-bottom:8px; }
  .box-list  { list-style:none; padding:0; font-size:12px; line-height:2; }
  .conditions-box {
    background:#FFF7ED; border:1px solid #F4A261; border-radius:8px;
    padding:14px 18px; margin:12px 0;
  }
  .audit-box {
    background:#F1F5F9; border:1px solid #CBD5E1; border-radius:8px;
    padding:14px 18px; font-size:12px; line-height:1.7; color:#374151; margin:12px 0;
  }
  .stress-grid { display:grid; grid-template-columns:repeat(3,1fr); gap:12px; margin:12px 0; }
  .stress-card { border-radius:8px; padding:14px; text-align:center; border:1px solid #E2E8F0; }
  .stress-label { font-size:11px; color:#6B7280; text-transform:uppercase;
                  letter-spacing:0.6px; margin-bottom:6px; }
  .stress-result { font-size:20px; font-weight:800; }
  .layer-grid { display:grid; grid-template-columns:repeat(3,1fr); gap:12px; margin:12px 0; }
  .layer-card { border-radius:8px; padding:14px; text-align:center; }
  .excel-badge {
    background:#D1FAE5; border:1px solid #10B981; border-radius:8px;
    padding:12px 18px; font-size:13px; color:#065F46; margin:12px 0;
    display:flex; align-items:center; gap:10px;
  }
  .new-assessment-btn {
    width:100%; padding:14px; background:#10B981; color:white; border:none;
    border-radius:10px; font-size:14px; font-weight:700; cursor:pointer;
    margin-top:16px; transition:opacity 0.2s;
  }
  .new-assessment-btn:hover { opacity:0.9; }
  .section-label {
    background:#0A1628; color:white; padding:8px 14px; border-radius:6px;
    font-size:12px; font-weight:600; margin:14px 0 10px;
  }
  @media (max-width:768px) {
    .form-row,.two-col,.metrics-grid,.stress-grid,.layer-grid { grid-template-columns:1fr; }
  }
</style>
</head>
<body>

<div class="hero">
  <h1>🏦 CREDIT RISK MODEL AI AGENT</h1><div style="margin-top:10px"><a href="/results" style="background:#F4A261;color:white;padding:6px 16px;border-radius:20px;font-size:12px;font-weight:600;text-decoration:none;">📊 View All Results</a></div>
  <p>Basel III · IFRS 9 · 3-Layer Ensemble · Powered by Claude Sonnet 4</p>
  <div class="tags">
    <span class="tag">Layer 1: Quantitative</span>
    <span class="tag">Layer 2: Behavioral</span>
    <span class="tag">Layer 3: Macro</span>
    <span class="tag">AI Meta-Aggregator</span>
  </div>
</div>

<div class="container">

  <div id="form-section">
    <div class="card">
      <div class="card-header">📋 Borrower Profile</div>
      <div class="card-body">

        <div class="toggle-row">
          <button class="toggle-btn active" onclick="setLoanType('personal')" id="btn-personal">
            👤 Personal / Consumer
          </button>
          <button class="toggle-btn" onclick="setLoanType('business')" id="btn-business">
            🏢 Business / Commercial
          </button>
        </div>

        <form id="borrowerForm">

          <div class="section-label">👤 Personal Information</div>
          <div class="form-row">
            <div class="form-group">
              <label>Full Name</label>
              <input type="text" name="name" placeholder="John Smith" required>
            </div>
            <div class="form-group">
              <label>Age</label>
              <input type="number" name="age" value="35" min="18" max="75">
            </div>
          </div>

          <div class="section-label">💰 Financial Profile</div>
          <div class="form-row">
            <div class="form-group">
              <label>Annual Income ($)</label>
              <input type="number" name="annual_income" value="85000">
            </div>
            <div class="form-group">
              <label>Monthly Debt Payments ($)</label>
              <input type="number" name="monthly_debt" value="1800">
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>Credit Score (300–850)</label>
              <input type="number" name="credit_score" value="680" min="300" max="850">
            </div>
            <div class="form-group">
              <label>Credit Utilisation (%)</label>
              <input type="number" name="credit_util" value="45" min="0" max="100">
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>Employment Type</label>
              <select name="employment_type">
                <option>Full-time Employed</option>
                <option>Self-employed</option>
                <option>Part-time</option>
                <option>Contract</option>
                <option>Retired</option>
                <option>Other</option>
              </select>
            </div>
            <div class="form-group">
              <label>Years in Current Job</label>
              <input type="number" name="employment_years" value="4.5" step="0.5">
            </div>
          </div>

          <div class="section-label">🏦 Loan Details</div>
          <div class="form-row">
            <div class="form-group">
              <label>Loan Amount ($)</label>
              <input type="number" name="loan_amount" value="120000">
            </div>
            <div class="form-group">
              <label>Loan Term (months)</label>
              <input type="number" name="loan_term" value="60">
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>Loan Purpose</label>
              <select name="loan_purpose">
                <option>Home Purchase</option>
                <option>Refinance</option>
                <option>Auto</option>
                <option>Business Expansion</option>
                <option>Debt Consolidation</option>
                <option>Working Capital</option>
                <option>Equipment Purchase</option>
                <option>Education</option>
                <option>Personal</option>
                <option>Other</option>
              </select>
            </div>
            <div class="form-group">
              <label>Facility Type</label>
              <select name="facility_type">
                <option>Term Loan</option>
                <option>Revolving Credit</option>
                <option>Line of Credit</option>
                <option>Overdraft</option>
                <option>Mortgage</option>
              </select>
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>Interest Rate (%)</label>
              <input type="number" name="interest_rate" value="6.5" step="0.1">
            </div>
            <div class="form-group">
              <label>Collateral Type</label>
              <select name="collateral_type">
                <option>Real Estate</option>
                <option>Equipment</option>
                <option>Receivables</option>
                <option>Inventory</option>
                <option>Government Bond</option>
                <option>Cash Deposit</option>
                <option>Unsecured</option>
              </select>
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>Collateral Value ($)</label>
              <input type="number" name="collateral_value" value="180000">
            </div>
            <div class="form-group">
              <label>Income Change YoY (%)</label>
              <input type="number" name="income_change" value="8" step="0.5">
            </div>
          </div>

          <div id="business-fields" style="display:none;">
            <div class="section-label">🏢 Business Details</div>
            <div class="form-row">
              <div class="form-group">
                <label>Company Name</label>
                <input type="text" name="company_name" placeholder="Acme Corp Ltd">
              </div>
              <div class="form-group">
                <label>Industry</label>
                <select name="industry">
                  <option>Technology</option>
                  <option>Healthcare</option>
                  <option>Real Estate</option>
                  <option>Manufacturing</option>
                  <option>Retail</option>
                  <option>Energy</option>
                  <option>Financial Services</option>
                  <option>Construction</option>
                  <option>Hospitality</option>
                  <option>Agriculture</option>
                  <option>Other</option>
                </select>
              </div>
            </div>
            <div class="form-row">
              <div class="form-group">
                <label>Business Age (years)</label>
                <input type="number" name="business_age" value="5" step="0.5">
              </div>
              <div class="form-group">
                <label>Number of Employees</label>
                <input type="number" name="num_employees" value="45">
              </div>
            </div>
            <div class="form-row">
              <div class="form-group">
                <label>Annual Revenue ($)</label>
                <input type="number" name="annual_revenue" value="2500000">
              </div>
              <div class="form-group">
                <label>EBITDA ($)</label>
                <input type="number" name="ebitda" value="420000">
              </div>
            </div>
            <div class="form-row">
              <div class="form-group">
                <label>Total Debt ($)</label>
                <input type="number" name="total_debt_biz" value="850000">
              </div>
              <div class="form-group">
                <label>Total Assets ($)</label>
                <input type="number" name="total_assets" value="2100000">
              </div>
            </div>
            <div class="form-row">
              <div class="form-group">
                <label>Current Assets ($)</label>
                <input type="number" name="current_assets" value="680000">
              </div>
              <div class="form-group">
                <label>Current Liabilities ($)</label>
                <input type="number" name="current_liab" value="320000">
              </div>
            </div>
          </div>

          <div class="section-label">📈 Behavioral History</div>
          <div class="form-row">
            <div class="form-group">
              <label>Late Payments (last 24 months)</label>
              <input type="number" name="late_payments" value="1" min="0" max="24">
            </div>
            <div class="form-group">
              <label>Hard Inquiries (last 12 months)</label>
              <input type="number" name="hard_inquiries" value="2" min="0" max="20">
            </div>
          </div>
          <div class="form-row">
            <div class="form-group">
              <label>New Accounts (last 12 months)</label>
              <input type="number" name="new_accounts" value="1" min="0" max="15">
            </div>
            <div class="form-group">
              <label>Credit Utilisation 6 Months Ago (%)</label>
              <input type="number" name="util_6m_ago" value="32" min="0" max="100">
            </div>
          </div>

          <button type="submit" class="submit-btn">
            🚀 Run Full Credit Assessment
          </button>
        </form>
      </div>
    </div>
  </div>

  <div class="progress" id="progress">
    <div class="progress-title">⚙️ Running Credit Assessment...</div>
    <div class="step" id="s1"><span class="step-icon">⚙️</span> Layer 1: Quantitative Analysis (PD/LGD/EAD/EL)</div>
    <div class="step" id="s2"><span class="step-icon">🧠</span> Layer 2: Behavioral Risk Engine</div>
    <div class="step" id="s3"><span class="step-icon">🌍</span> Layer 3: Macro Overlay</div>
    <div class="step" id="s4"><span class="step-icon">🤖</span> Claude AI Meta-Aggregator</div>
    <div class="step" id="s5"><span class="step-icon">📊</span> Generating Report & Saving to Excel</div>
  </div>

  <div id="results"></div>

</div>

<script>
let loanType = 'personal';

function setLoanType(type) {
  loanType = type;
  document.getElementById('btn-personal').classList.toggle('active', type==='personal');
  document.getElementById('btn-business').classList.toggle('active', type==='business');
  document.getElementById('business-fields').style.display = type==='business' ? 'block' : 'none';
}

function setStep(n, status) {
  const el = document.getElementById('s'+n);
  el.className = 'step ' + status;
  if (status==='done')   el.querySelector('.step-icon').textContent = '✅';
  if (status==='error')  el.querySelector('.step-icon').textContent = '❌';
}

function gradeColor(g) {
  return {AAA:'#10B981',AA:'#10B981',A:'#34D399',BBB:'#2E86AB',
          BB:'#F4A261',B:'#F59E0B',CCC:'#EF4444',D:'#7F1D1D'}[g]||'#6B7280';
}
function decisionColor(d) {
  return {'Auto Approve':'#10B981','Approve':'#34D399','Conditional Approval':'#F4A261',
          'Manual Review':'#F59E0B','Decline':'#EF4444'}[d]||'#6B7280';
}
function stressBg(v)   { return v==='Pass'?'#F0FDF4':'#FEF2F2'; }
function stressClr(v)  { return v==='Pass'?'#065F46':'#991B1B'; }

document.getElementById('borrowerForm').addEventListener('submit', async function(e) {
  e.preventDefault();
  const data = Object.fromEntries(new FormData(e.target));
  data.loan_type = loanType;

  document.getElementById('progress').style.display = 'block';
  document.getElementById('results').style.display  = 'none';
  document.getElementById('results').innerHTML      = '';
  document.querySelector('.submit-btn').disabled    = true;
  [1,2,3,4,5].forEach(n => { document.getElementById('s'+n).className='step'; });

  setStep(1,'active');
  setTimeout(()=>{ setStep(1,'done'); setStep(2,'active'); }, 1000);
  setTimeout(()=>{ setStep(2,'done'); setStep(3,'active'); }, 2000);
  setTimeout(()=>{ setStep(3,'done'); setStep(4,'active'); }, 3500);
  setTimeout(()=>{ setStep(4,'done'); setStep(5,'active'); }, 6000);

  try {
    const res    = await fetch('/assess', {
      method:'POST', headers:{'Content-Type':'application/json'},
      body: JSON.stringify(data)
    });
    const r = await res.json();

    setStep(5,'done');
    document.getElementById('progress').style.display = 'none';
    document.querySelector('.submit-btn').disabled = false;

    if (r.error) {
      document.getElementById('results').style.display = 'block';
      document.getElementById('results').innerHTML =
        `<div style="background:#FEE2E2;border-radius:8px;padding:16px;color:#991B1B;">❌ Error: ${r.error}</div>`;
      return;
    }

    const gc     = gradeColor(r.grade);
    const dc     = decisionColor(r.decision);
    const stress = r.stress_test || {};
    const conds  = (r.conditions||[]).map(c=>`<li>📋 ${c}</li>`).join('');
    const strs   = (r.strengths||[]).map(s=>`<li style="color:#065F46;">✅ ${s}</li>`).join('');
    const flags  = (r.red_flags||[]).map(f=>`<li style="color:#991B1B;">🚩 ${f}</li>`).join('');

    document.getElementById('results').style.display = 'block';
    document.getElementById('results').innerHTML = `

      <div class="verdict">
        <div class="verdict-label">CREDIT COMMITTEE VERDICT</div>
        <div class="verdict-score" style="color:${gc};">${r.composite_score}</div>
        <div class="verdict-sub">out of 1000</div>
        <div class="verdict-pills">
          <span class="pill" style="background:${gc};">${r.grade}</span>
          <span class="pill" style="background:${dc};">${r.decision}</span>
          <span class="pill" style="background:rgba(255,255,255,0.15);">Confidence: ${r.confidence}%</span>
        </div>
      </div>

      <div class="metrics-grid">
        <div class="metric" style="border-top-color:${gc};">
          <div class="metric-label">Adjusted PD</div>
          <div class="metric-value">${(r.adjusted_pd*100).toFixed(2)}%</div>
        </div>
        <div class="metric" style="border-top-color:#2E86AB;">
          <div class="metric-label">Expected Loss</div>
          <div class="metric-value">$${Number(r.el).toLocaleString()}</div>
        </div>
        <div class="metric" style="border-top-color:#F4A261;">
          <div class="metric-label">Suggested Rate</div>
          <div class="metric-value">${Number(r.suggested_rate).toFixed(2)}%</div>
        </div>
        <div class="metric" style="border-top-color:#8B5CF6;">
          <div class="metric-label">IFRS 9 Stage</div>
          <div class="metric-value" style="font-size:14px;">${r.ifrs9_stage}</div>
        </div>
      </div>

      <div class="card">
        <div class="card-header">📊 Layer Score Breakdown</div>
        <div class="card-body">
          <div class="layer-grid">
            <div class="layer-card" style="background:#EEF2FF;">
              <div style="font-size:11px;color:#6B7280;text-transform:uppercase;margin-bottom:4px;">Layer 1 · Quantitative</div>
              <div style="font-size:28px;font-weight:800;color:#2E86AB;">${r.l1_score}</div>
              <div style="font-size:11px;color:#6B7280;">/ 1000</div>
            </div>
            <div class="layer-card" style="background:#F0FDF4;">
              <div style="font-size:11px;color:#6B7280;text-transform:uppercase;margin-bottom:4px;">Layer 2 · Behavioral</div>
              <div style="font-size:28px;font-weight:800;color:#10B981;">${r.l2_score}</div>
              <div style="font-size:11px;color:#6B7280;">/ 1000</div>
            </div>
            <div class="layer-card" style="background:#FFF7ED;">
              <div style="font-size:11px;color:#6B7280;text-transform:uppercase;margin-bottom:4px;">Layer 3 · Macro</div>
              <div style="font-size:28px;font-weight:800;color:#F4A261;">${r.l3_score}</div>
              <div style="font-size:11px;color:#6B7280;">/ 1000</div>
            </div>
          </div>
        </div>
      </div>

      <div class="card">
        <div class="card-header">📝 Credit Committee Rationale</div>
        <div class="card-body">
          <div class="rationale-box">${r.rationale}</div>
        </div>
      </div>

      <div class="card">
        <div class="card-header">🔍 AI Reasoning Chain</div>
        <div class="card-body">
          <div class="reasoning-box">
[Layer Consistency]  ${r.layer_consistency}<br>
[Dominant Risk]      ${r.dominant_risk_driver}<br>
[AI Adjustment]      ${r.ai_score_adjustment>0?'+':''}${r.ai_score_adjustment} pts — ${r.ai_adjustment_reason}<br>
[IFRS 9 Stage]       ${r.ifrs9_stage}<br>
[Final Decision]     ${r.decision}
          </div>
        </div>
      </div>

      <div class="card">
        <div class="card-header">🧪 Stress Test Results</div>
        <div class="card-body">
          <div class="stress-grid">
            <div class="stress-card" style="background:${stressBg(stress.mild_recession)};">
              <div class="stress-label">Mild Recession</div>
              <div class="stress-result" style="color:${stressClr(stress.mild_recession)};">${stress.mild_recession||'N/A'}</div>
            </div>
            <div class="stress-card" style="background:${stressBg(stress.moderate_recession)};">
              <div class="stress-label">Moderate Recession</div>
              <div class="stress-result" style="color:${stressClr(stress.moderate_recession)};">${stress.moderate_recession||'N/A'}</div>
            </div>
            <div class="stress-card" style="background:${stressBg(stress.severe_recession)};">
              <div class="stress-label">Severe Recession</div>
              <div class="stress-result" style="color:${stressClr(stress.severe_recession)};">${stress.severe_recession||'N/A'}</div>
            </div>
          </div>
        </div>
      </div>

      <div class="two-col">
        <div class="strengths-box">
          <div class="box-title" style="color:#065F46;">💪 Strengths</div>
          <ul class="box-list">${strs||'<li style="color:#6B7280">None noted</li>'}</ul>
        </div>
        <div class="flags-box">
          <div class="box-title" style="color:#991B1B;">🚩 Red Flags</div>
          <ul class="box-list">${flags||'<li style="color:#6B7280">None identified</li>'}</ul>
        </div>
      </div>

      ${conds?`
      <div class="conditions-box">
        <div style="font-weight:700;color:#92400E;margin-bottom:8px;">📋 Approval Conditions</div>
        <ul style="padding-left:16px;font-size:13px;color:#92400E;line-height:1.8;">${conds}</ul>
      </div>`:''}

      <div class="card">
        <div class="card-header">🔍 Full Audit Trail</div>
        <div class="card-body">
          <div class="audit-box">${r.audit_trail}</div>
        </div>
      </div>

      <div class="excel-badge">
        📊 Saved to Google Sheets — Record #${r.total_records} · <a href='https://docs.google.com/spreadsheets/d/13pFG_57Eb6EOGeuw90B65Z6q2LkuhSqfAFLPYE83zrk' target='_blank' style='color:#065F46;font-weight:700'>Open Sheet ↗</a>
      </div>

      <button class="new-assessment-btn" onclick="newAssessment()">
        🔄 Run New Assessment
      </button>
    `;

    document.getElementById('results').scrollIntoView({behavior:'smooth'});

  } catch(err) {
    setStep(5,'error');
    document.querySelector('.submit-btn').disabled = false;
    document.getElementById('results').style.display = 'block';
    document.getElementById('results').innerHTML =
      `<div style="background:#FEE2E2;border-radius:8px;padding:16px;color:#991B1B;">❌ Error: ${err.message}</div>`;
  }
});

function newAssessment() {
  document.getElementById('results').style.display  = 'none';
  document.getElementById('results').innerHTML      = '';
  document.getElementById('progress').style.display = 'none';
  window.scrollTo({top:0,behavior:'smooth'});
  document.getElementById('borrowerForm').reset();
}
</script>
</body>
</html>
"""

# ============================================================
# CREDIT RISK ENGINE
# ============================================================


# In-memory results storage
ASSESSMENTS = []

def logistic(x):
    return 1 / (1 + math.exp(-x))

def run_layer1(bw):
    th  = CONFIG["thresholds"]
    ads = bw["monthly_debt"] * 12
    ebit = bw["ebitda"] * 0.85
    ie   = bw["loan_amount"] * bw["interest_rate"]

    dscr  = bw["ebitda"] / max(ads, 1)
    lev   = bw["total_debt"] / max(bw["total_assets"], 1)
    icr   = ebit / max(ie, 1)
    cr    = bw["current_assets"] / max(bw["current_liab"], 1)
    qr    = (bw["current_assets"] - bw.get("receivables",0)*0.3) / max(bw["current_liab"],1)
    dti   = bw["dti_ratio"]
    ltv   = bw["ltv_ratio"]

    z = (2.5*(dscr-th["dscr_min"]) - 3.0*(lev-th["leverage_max"]) +
         1.5*(icr-th["interest_cov_min"]) + 1.2*(cr-th["current_ratio_min"]) -
         2.0*(dti-th["dti_max"]) + 0.005*(bw["credit_score"]-600) +
         0.15*min(bw["employment_years"],10))

    spd  = CONFIG["sector_base_pd"].get(bw["industry"], 0.030)
    pd   = max(0.001, min(0.999, (1-logistic(z))*0.40 + spd*0.60))

    if pd<0.005:   grade="AAA"
    elif pd<0.010: grade="AA"
    elif pd<0.020: grade="A"
    elif pd<0.050: grade="BBB"
    elif pd<0.100: grade="BB"
    elif pd<0.200: grade="B"
    elif pd<0.300: grade="CCC"
    else:          grade="D"

    lgd  = min(0.95, CONFIG["lgd_by_collateral"].get(bw["collateral_type"],0.90) + max(0,(ltv-0.60)*0.3))
    ead  = bw["loan_amount"] * CONFIG["ccf_by_type"].get(bw["facility_type"],1.0)
    el   = pd * lgd * ead
    cs   = pd * lgd * 1.5
    sr   = 0.0525 + cs

    s_dscr = min(100,max(0,(dscr/2.5)*100))
    s_lev  = min(100,max(0,(1-lev)*100))
    s_icr  = min(100,max(0,(icr/8)*100))
    s_cr   = min(100,max(0,(cr/3)*100))
    s_pd   = min(100,max(0,(1-pd*5)*100))
    s_cred = min(100,max(0,(bw["credit_score"]-300)/5.5))
    s_dti  = min(100,max(0,(1-dti)*100))

    l1s = round((s_dscr*0.20+s_lev*0.15+s_icr*0.15+s_cr*0.10+s_pd*0.20+s_cred*0.10+s_dti*0.10)*10)
    ifrs = "Stage 1" if pd<0.02 else "Stage 2" if pd<0.15 else "Stage 3"

    return {"dscr":round(dscr,3),"leverage_ratio":round(lev,3),"interest_cov":round(icr,2),
            "current_ratio":round(cr,2),"quick_ratio":round(qr,2),"dti":round(dti,3),
            "ltv":round(ltv,3),"pd":round(pd,4),"pd_grade":grade,"lgd":round(lgd,3),
            "recovery_rate":round(1-lgd,3),"ead":round(ead,2),"el":round(el,2),
            "credit_spread_bps":round(cs*10000,1),"suggested_rate":round(sr,4),
            "ifrs9_stage":ifrs,"l1_score":l1s}

def run_layer2(bw):
    th = CONFIG["thresholds"]
    flags, scores = [], {}

    un  = bw["credit_util_pct"]
    ud  = un - bw["util_6m_ago_pct"]
    if un>th["utilisation_max"]: flags.append({"flag":f"Utilisation {un:.0%} exceeds 80%","severity":"HIGH"})
    elif un>0.60: flags.append({"flag":f"Utilisation {un:.0%} approaching danger zone","severity":"MEDIUM"})
    if ud>0.15: flags.append({"flag":f"Utilisation rising +{ud:.0%} in 6 months","severity":"HIGH"})
    scores["Utilisation"] = round(max(0,100-(un*80)-(max(0,ud)*60)),1)

    late = bw["late_payments"]
    ps   = 100 if late==0 else 78 if late<=1 else 55 if late<=3 else 20
    st   = "Perfect" if late==0 else "Minor" if late<=1 else "Concerning" if late<=3 else "High Risk"
    if late>0: flags.append({"flag":f"{late} late payment(s) in 24 months",
                              "severity":"LOW" if late<=1 else "MEDIUM" if late<=3 else "HIGH"})
    scores["Payment History"] = ps

    inq = bw["hard_inquiries"]
    iq  = 95 if inq<=1 else 75 if inq<=3 else 50 if inq<=6 else 20
    if inq>3: flags.append({"flag":f"{inq} hard inquiries in 12 months","severity":"MEDIUM" if inq<=6 else "HIGH"})
    scores["Inquiries"] = iq

    na = bw["new_accounts"]
    aq = 95 if na<=1 else 75 if na<=3 else 55 if na<=5 else 30
    if na>5: flags.append({"flag":f"{na} new accounts in 12 months","severity":"MEDIUM"})
    scores["New Accounts"] = aq

    ic = bw["income_change_pct"]
    if ic>0.40:    is_=40; flags.append({"flag":f"Income spike +{ic:.0%}","severity":"HIGH"})
    elif ic>0.25:  is_=65; flags.append({"flag":f"Income +{ic:.0%} — verify","severity":"MEDIUM"})
    elif ic<-0.20: is_=30; flags.append({"flag":f"Income decline {ic:.0%}","severity":"HIGH"})
    elif ic<-0.10: is_=60; flags.append({"flag":f"Income declining {ic:.0%}","severity":"MEDIUM"})
    else:          is_=90
    scores["Income Stability"] = is_

    et = bw["employment_type"]
    eb = 85 if et=="Full-time Employed" else 65 if et in ["Self-employed","Contract"] else 55 if et=="Part-time" else 50
    if et!="Full-time Employed": flags.append({"flag":f"Employment: {et}","severity":"LOW"})
    tb = min(20, bw["employment_years"]*2)
    if bw["employment_years"]<1:
        flags.append({"flag":"Less than 1 year in role","severity":"MEDIUM"})
        tb=-10
    scores["Employment"] = min(100,max(0,eb+tb))

    hf = len([f for f in flags if f["severity"]=="HIGH"])
    mf = len([f for f in flags if f["severity"]=="MEDIUM"])
    lf = len([f for f in flags if f["severity"]=="LOW"])
    pen= hf*15+mf*7+lf*3
    l2s= round(max(0,min(1000,(sum(scores.values())/len(scores)-pen)*10)))

    return {"scores":scores,"flags":flags,"high_flags":hf,"medium_flags":mf,
            "low_flags":lf,"anomaly_penalty":pen,"l2_score":l2s,"payment_status":st}

def run_layer3(bw, l1):
    industry = bw.get("industry","Other")
    # Use calibrated baseline values (avoids token-heavy web search)
    try:
        resp = client.messages.create(
            model=MODEL, max_tokens=400,
            messages=[{"role":"user","content":
                f"Return ONLY a JSON object with realistic current US macro estimates for {industry} sector. "
                f"No explanation. Use these approximate values as baseline and adjust for {industry}: "
                f'{{"fed_funds_rate":4.33,"treasury_10y":4.45,"treasury_2y":4.92,"yield_spread":-0.47,'
                f'"cpi_inflation":3.2,"unemployment_rate":4.1,"gdp_growth_qoq":2.4,"vix_index":18.4,'
                f'"economic_cycle":"Contraction","sector_default_rate":3.0,"sector_outlook":"Neutral",'
                f'"sector_notes":"{industry} sector mixed signals.","macro_risk_score":58,'
                f'"macro_summary":"Elevated rates with softening growth. Labour market resilient."}}'}]
        )
        text  = resp.content[0].text.strip()
        match = re.search(r'\{[\s\S]*\}', text)
        macro = json.loads(match.group()) if match else None
    except:
        macro = None

    if not macro:
        macro = {"fed_funds_rate":4.33,"treasury_10y":4.45,"treasury_2y":4.92,
                 "yield_spread":-0.47,"cpi_inflation":3.2,"unemployment_rate":4.1,
                 "gdp_growth_qoq":2.4,"vix_index":18.4,"economic_cycle":"Contraction",
                 "sector_default_rate":CONFIG["sector_base_pd"].get(industry,0.030)*100,
                 "sector_outlook":"Neutral","sector_notes":f"{industry} mixed signals.",
                 "macro_risk_score":58,"macro_summary":"Elevated rates with softening growth."}

    cycle = macro.get("economic_cycle","Contraction")
    cadj  = {"Expansion":-0.005,"Peak":0.005,"Contraction":0.015,"Trough":0.025}.get(cycle,0.010)
    vix   = macro.get("vix_index",20)
    vadj  = max(0,(vix-20)*0.001)
    sp    = macro.get("yield_spread",0)
    curv  = max(0,-sp*0.005) if sp<0 else 0
    tadj  = cadj+vadj+curv
    apd   = min(0.99, l1["pd"]+tadj)
    om    = {"Positive":10,"Neutral":0,"Cautious":-10,"Negative":-20}.get(macro.get("sector_outlook","Neutral"),0)
    l3s   = round(min(1000,max(0,(macro.get("macro_risk_score",60)+om)*10)))

    return {"macro_data":macro,"cycle":cycle,"total_pd_adj":round(tadj,4),
            "adjusted_pd":round(apd,4),"l3_score":l3s}

def run_ai(bw, l1, l2, l3):
    raw = round(l1["l1_score"]*0.40 + l2["l2_score"]*0.35 + l3["l3_score"]*0.25)

    prompt = f"""Credit assessment for {bw['name']}, ${bw['loan_amount']:,.0f} {bw['facility_type']}.
L1:{l1['l1_score']} PD:{l1['pd']:.2%} DSCR:{l1['dscr']} DTI:{l1['dti']:.1%} LTV:{l1['ltv']:.1%} EL:${l1['el']:,.0f}
L2:{l2['l2_score']} Flags:{l2['high_flags']}H·{l2['medium_flags']}M·{l2['low_flags']}L
L3:{l3['l3_score']} Cycle:{l3['cycle']} AdjPD:{l3['adjusted_pd']:.2%}
Raw:{raw}/1000 IFRS9:{l1['ifrs9_stage']}

Return ONLY this JSON (no markdown):
{{"composite_score":<int>,"grade":<AAA|AA|A|BBB|BB|B|CCC|D>,
"decision":<Auto Approve|Approve|Conditional Approval|Manual Review|Decline>,
"confidence":<int 70-99>,"ifrs9_stage":<Stage 1|Stage 2|Stage 3>,
"adjusted_pd":<float>,"suggested_rate":<float %>,"layer_consistency":<str>,
"dominant_risk_driver":<str>,
"stress_test":{{"mild_recession":<Pass|Fail>,"moderate_recession":<Pass|Fail>,"severe_recession":<Pass|Fail>}},
"regulatory_flags":[<str>],"red_flags":[<str>],"conditions":[<str>],
"strengths":[<str>],"rationale":"<4 sentences>","audit_trail":"<step by step>",
"ai_score_adjustment":<int -50 to 50>,"ai_adjustment_reason":"<str>"}}"""

    resp = client.messages.create(
        model=MODEL, max_tokens=2000,
        system="You are a Senior Credit Risk Officer. Return ONLY valid JSON. Be concise.",
        messages=[{"role":"user","content":prompt}]
    )
    text = resp.content[0].text.strip()
    # Remove markdown code blocks if present
    text = re.sub(r"```json|```", "", text).strip()
    print(f"AI RAW: {text[:300]}", flush=True)
    try:
        match = re.search(r'\{[\s\S]*\}', text)
        if match:
            return json.loads(match.group())
    except Exception as e:
        print(f"AI JSON error: {e}", flush=True)
        print(f"AI response: {text[:200]}", flush=True)
    return {}


def save_to_sheets(bw,l1,l2,l3,ai):
    total_records=0
    try:
        import psycopg2
        db_url=os.environ.get("DATABASE_URL","")
        if not db_url:
            print("No DATABASE_URL found", flush=True)
            return 0
        conn=psycopg2.connect(db_url, sslmode="require")
        cur=conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS assessments (
                id SERIAL PRIMARY KEY,
                timestamp TEXT,
                borrower TEXT,
                loan_amount FLOAT,
                grade TEXT,
                score INTEGER,
                decision TEXT,
                pd TEXT,
                lgd TEXT,
                el FLOAT,
                suggested_rate FLOAT,
                ifrs9 TEXT,
                l1_score INTEGER,
                l2_score INTEGER,
                l3_score INTEGER,
                confidence INTEGER,
                stress_mild TEXT,
                stress_moderate TEXT,
                stress_severe TEXT,
                cycle TEXT,
                rationale TEXT,
                conditions TEXT,
                red_flags TEXT
            )
        """)
        stress=ai.get("stress_test",{})
        cur.execute("""
            INSERT INTO assessments
            (timestamp,borrower,loan_amount,grade,score,decision,pd,lgd,el,
             suggested_rate,ifrs9,l1_score,l2_score,l3_score,confidence,
             stress_mild,stress_moderate,stress_severe,cycle,rationale,conditions,red_flags)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """, (
            datetime.now().strftime("%d %b %Y %H:%M"),
            bw["name"],bw["loan_amount"],
            ai.get("grade",""),ai.get("composite_score",0),ai.get("decision",""),
            f"{l3['adjusted_pd']:.2%}",f"{l1['lgd']:.1%}",round(l1["el"],0),
            ai.get("suggested_rate",0),ai.get("ifrs9_stage",""),
            l1["l1_score"],l2["l2_score"],l3["l3_score"],ai.get("confidence",0),
            stress.get("mild_recession",""),stress.get("moderate_recession",""),
            stress.get("severe_recession",""),l3["cycle"],
            ai.get("rationale",""),
            " | ".join(ai.get("conditions",[])),
            " | ".join(ai.get("red_flags",[])),
        ))
        conn.commit()
        cur.execute("SELECT COUNT(*) FROM assessments")
        total_records=cur.fetchone()[0]
        cur.close()
        conn.close()
        print(f"Saved to Supabase — Record #{total_records}", flush=True)
    except Exception as e:
        import traceback
        print(f"Database error: {e}", flush=True)
        traceback.print_exc()
    # Always save to memory regardless of database
    ASSESSMENTS.append({
        "timestamp": datetime.now().strftime("%d %b %Y %H:%M"),
        "borrower": bw["name"],
        "loan": bw["loan_amount"],
        "grade": ai.get("grade",""),
        "score": ai.get("composite_score",0),
        "decision": ai.get("decision",""),
        "pd": f"{l3['adjusted_pd']:.2%}",
        "el": round(l1["el"],0),
        "rate": ai.get("suggested_rate",0),
        "ifrs9": ai.get("ifrs9_stage",""),
        "rationale": ai.get("rationale",""),
    })
    return total_records

def save_excel(bw, l1, l2, l3, ai):
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        wb = openpyxl.Workbook()

    ws = wb["Portfolio"] if "Portfolio" in wb.sheetnames else wb.active
    ws.title = "Portfolio"

    headers = ["Timestamp","Borrower","Loan ($)","Grade","Score","Decision",
               "PD","LGD","EL ($)","Rate (%)","IFRS 9",
               "L1","L2","L3","Confidence",
               "Stress Mild","Stress Moderate","Stress Severe",
               "Cycle","Rationale","Conditions","Red Flags"]

    if ws.max_row==1 and ws["A1"].value is None:
        for c,h in enumerate(headers,1):
            cell = ws.cell(row=1,column=c,value=h)
            cell.fill = PatternFill("solid",fgColor="0A1628")
            cell.font = Font(bold=True,color="FFFFFF",size=10)
            cell.alignment = Alignment(horizontal="center",vertical="center")
            ws.column_dimensions[get_column_letter(c)].width = max(14,len(h)+2)
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

    stress = ai.get("stress_test",{})
    nr  = ws.max_row + 1
    row = [
        datetime.now().strftime("%d %b %Y %H:%M"),
        bw["name"], bw["loan_amount"],
        ai.get("grade",""), ai.get("composite_score",0), ai.get("decision",""),
        f"{l3['adjusted_pd']:.2%}", f"{l1['lgd']:.1%}", round(l1["el"],0),
        ai.get("suggested_rate",0), ai.get("ifrs9_stage",""),
        l1["l1_score"], l2["l2_score"], l3["l3_score"], ai.get("confidence",0),
        stress.get("mild_recession",""), stress.get("moderate_recession",""), stress.get("severe_recession",""),
        l3["cycle"], ai.get("rationale",""),
        " | ".join(ai.get("conditions",[])), " | ".join(ai.get("red_flags",[])),
    ]

    gc = {"AAA":"10B981","AA":"10B981","A":"34D399","BBB":"2E86AB",
          "BB":"F4A261","B":"F59E0B","CCC":"EF4444","D":"7F1D1D"}
    dc = {"Auto Approve":"10B981","Approve":"34D399","Conditional Approval":"F4A261",
          "Manual Review":"F59E0B","Decline":"EF4444"}

    for c,val in enumerate(row,1):
        cell = ws.cell(row=nr,column=c,value=val)
        cell.alignment = Alignment(horizontal="center",vertical="center")
        cell.font = Font(size=10)
        if nr%2==0: cell.fill = PatternFill("solid",fgColor="F1F5F9")

    g = ai.get("grade","")
    if g in gc:
        ws.cell(row=nr,column=4).fill = PatternFill("solid",fgColor=gc[g])
        ws.cell(row=nr,column=4).font = Font(bold=True,color="FFFFFF",size=11)

    d = ai.get("decision","")
    if d in dc:
        ws.cell(row=nr,column=6).fill = PatternFill("solid",fgColor=dc[d])
        ws.cell(row=nr,column=6).font = Font(bold=True,color="FFFFFF",size=10)

    ws.row_dimensions[nr].height = 20
    wb.save(EXCEL_PATH)
    return ws.max_row - 1

# ============================================================
# ROUTES
# ============================================================

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/assess", methods=["POST"])
def assess():
    try:
        d      = request.json
        is_biz = d.get("loan_type") == "business"
        income = float(d.get("annual_income",85000))
        debt   = float(d.get("monthly_debt",1800))
        loan   = float(d.get("loan_amount",120000))
        colval = float(d.get("collateral_value",180000))

        bw = {
            "name":             d.get("name","Unknown"),
            "age":              int(d.get("age",35)),
            "loan_type":        "Business" if is_biz else "Personal",
            "is_business":      is_biz,
            "annual_income":    income,
            "monthly_debt":     debt,
            "credit_score":     int(d.get("credit_score",680)),
            "credit_util_pct":  float(d.get("credit_util",45))/100,
            "employment_type":  d.get("employment_type","Full-time Employed"),
            "employment_years": float(d.get("employment_years",4.5)),
            "loan_amount":      loan,
            "loan_term_months": int(d.get("loan_term",60)),
            "loan_purpose":     d.get("loan_purpose","Personal"),
            "facility_type":    d.get("facility_type","Term Loan"),
            "interest_rate":    float(d.get("interest_rate",6.5))/100,
            "collateral_type":  d.get("collateral_type","Real Estate"),
            "collateral_value": colval,
            "industry":         d.get("industry","Other") if is_biz else "Other",
            "company_name":     d.get("company_name","") if is_biz else "",
            "business_age":     float(d.get("business_age",0)) if is_biz else 0,
            "annual_revenue":   float(d.get("annual_revenue",income)) if is_biz else income,
            "ebitda":           float(d.get("ebitda",income*0.3)) if is_biz else income*0.3,
            "total_debt":       float(d.get("total_debt_biz",debt*12)) if is_biz else debt*12,
            "total_assets":     float(d.get("total_assets",colval)) if is_biz else colval,
            "current_assets":   float(d.get("current_assets",income*0.5)) if is_biz else income*0.5,
            "current_liab":     float(d.get("current_liab",debt*6)) if is_biz else debt*6,
            "receivables":      float(d.get("receivables",0)),
            "num_employees":    int(d.get("num_employees",1)),
            "util_6m_ago_pct":  float(d.get("util_6m_ago",32))/100,
            "late_payments":    int(d.get("late_payments",0)),
            "hard_inquiries":   int(d.get("hard_inquiries",0)),
            "new_accounts":     int(d.get("new_accounts",0)),
            "income_change_pct":float(d.get("income_change",0))/100,
            "monthly_income":   income/12,
            "dti_ratio":        debt/(income/12),
            "ltv_ratio":        loan/max(colval,1),
        }

        l1 = run_layer1(bw)
        l2 = run_layer2(bw)
        l3 = run_layer3(bw, l1)
        ai = run_ai(bw, l1, l2, l3)

        total = save_excel(bw, l1, l2, l3, ai)

        return jsonify({
            **ai,
            "l1_score":     l1["l1_score"],
            "l2_score":     l2["l2_score"],
            "l3_score":     l3["l3_score"],
            "el":           round(l1["el"],0),
            "adjusted_pd":  l3["adjusted_pd"],
            "ifrs9_stage":  l1["ifrs9_stage"],
            "total_records":total,
        })

    except Exception as e:
        return jsonify({"error": str(e)})

# ============================================================
# RUN
# ============================================================
if __name__ == "__main__":
    print("🏦 Credit Risk Model starting...")
    print("📂 Open your browser at: http://127.0.0.1:8080")
    app.run(debug=False, port=8080, host="127.0.0.1")

@app.route("/results")
def results():
    rows = ""
    for a in reversed(ASSESSMENTS):
        grade_color = {"AAA":"#10B981","AA":"#10B981","A":"#34D399","BBB":"#2E86AB","BB":"#F4A261","B":"#F59E0B","CCC":"#EF4444","D":"#7F1D1D"}.get(a["grade"],"#6B7280")
        dec_color = {"Auto Approve":"#10B981","Approve":"#34D399","Conditional Approval":"#F4A261","Manual Review":"#F59E0B","Decline":"#EF4444"}.get(a["decision"],"#6B7280")
        rows += f"""<tr style="border-bottom:1px solid #F1F5F9;">
            <td style="padding:10px 12px;font-size:13px;">{a["timestamp"]}</td>
            <td style="padding:10px 12px;font-size:13px;font-weight:600;">{a["borrower"]}</td>
            <td style="padding:10px 12px;font-size:13px;">${a["loan"]:,.0f}</td>
            <td style="padding:10px 12px;text-align:center;"><span style="background:{grade_color};color:white;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:700;">{a["grade"]}</span></td>
            <td style="padding:10px 12px;text-align:center;font-weight:700;color:{grade_color};">{a["score"]}</td>
            <td style="padding:10px 12px;text-align:center;"><span style="background:{dec_color};color:white;padding:3px 10px;border-radius:12px;font-size:11px;font-weight:600;">{a["decision"]}</span></td>
            <td style="padding:10px 12px;text-align:center;">{a["pd"]}</td>
            <td style="padding:10px 12px;text-align:center;">${a["el"]:,.0f}</td>
            <td style="padding:10px 12px;text-align:center;">{a["rate"]:.2f}%</td>
            <td style="padding:10px 12px;text-align:center;">{a["ifrs9"]}</td>
        </tr>"""
    html = f"""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>Credit Risk Portfolio</title>
<style>*{{box-sizing:border-box;margin:0;padding:0}}body{{font-family:'Segoe UI',sans-serif;background:#F8FAFC}}
.hero{{background:linear-gradient(135deg,#0A1628,#1A3A5C);padding:24px 32px;color:white}}
.hero h1{{font-size:22px;font-weight:800}}.hero p{{color:#F4A261;font-size:12px;margin-top:4px}}
.container{{max-width:1200px;margin:24px auto;padding:0 20px}}
.card{{background:white;border:1px solid #E2E8F0;border-radius:12px;overflow:hidden}}
.ch{{background:#0A1628;color:white;padding:12px 18px;font-weight:700;font-size:13px;display:flex;justify-content:space-between;align-items:center}}
table{{width:100%;border-collapse:collapse}}
th{{background:#1A3A5C;color:white;padding:10px 12px;font-size:11px;text-transform:uppercase;letter-spacing:.6px;text-align:left}}
tr:hover{{background:#F8FAFC}}.btn{{background:#2E86AB;color:white;padding:8px 18px;border-radius:8px;text-decoration:none;font-size:13px;font-weight:600}}
.empty{{text-align:center;padding:40px;color:#6B7280;font-size:14px}}</style></head>
<body><div class="hero"><h1>🏦 Credit Risk Portfolio</h1>
<p>Session assessments · {len(ASSESSMENTS)} total records</p></div>
<div class="container"><div class="card">
<div class="ch"><span>📊 Assessment History</span><a href="/" class="btn">+ New Assessment</a></div>
<table><tr><th>Time</th><th>Borrower</th><th>Loan</th><th>Grade</th><th>Score</th><th>Decision</th><th>PD</th><th>Exp Loss</th><th>Rate</th><th>IFRS9</th></tr>
{"".join(rows) if ASSESSMENTS else "<tr><td colspan=10 class=\'empty\'>No assessments yet. <a href=\'/\'>Run your first →</a></td></tr>"}
</table></div></div></body></html>"""
    return html

# updated Wed Jun  3 07:32:55 EDT 2026
